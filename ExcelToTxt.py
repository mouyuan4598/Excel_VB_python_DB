from openpyxl import load_workbook
from datetime import datetime
from datetime import date
wb = load_workbook('./Excel/ExpensesTable_ver 2.1_YST_Final_9.23.18.xlsm')
ws = wb['Expense']

f = open("./data/data.txt", "r")
linelist = f.readlines()
f.close()
i = 0
a=''
while linelist[-1][i] != ',':
    a = a + linelist[-1][i]
    i = i + 1

end = ws['A1'].value + 1
gap = 8-3
index = int(a) + 1
data = []



while index<end:
    d = datetime.date(ws['C' + str(index+gap)].value)
    data.append([ws['B' + str(index+gap)].value, d.isoformat(),ws['E' + str(index+gap)].value,ws['F' + str(index+gap)].value,ws['G' + str(index+gap)].value,ws['H' + str(index+gap)].value,ws['I' + str(index+gap)].value,ws['J' + str(index+gap)].value,ws['K' + str(index+gap)].value,ws['L' + str(index+gap)].value])
    index = index +1
    

f = open("./data/data.txt", "a")
for i in range(len(data)):
    for j in range(len(data[i])):
        f.write(str(data[i][j]))
        f.write(',')
    f.write('\n')
f.close()